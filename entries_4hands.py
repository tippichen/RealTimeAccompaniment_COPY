import os
import io
import shutil
import sys

# Attempt to load openpyxl, provide error message if not found
try:
    import openpyxl
except ImportError:
    print("Error: 'openpyxl' package not found. Please run 'pip install openpyxl' to install it.")
    sys.exit(1)

# Resolve encoding issues to ensure proper output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Set relative paths
script_dir = os.path.dirname(os.path.abspath(__file__))
BASE_DATA_DIR = os.path.join(script_dir, "data_management2026")

# File to read: database_entry.xlsx
DB_PATH = os.path.join(BASE_DATA_DIR, "database_entry.xlsx")

# Output directory: interpretation_by_entry
ENTRY_DIR = os.path.join(BASE_DATA_DIR, "interpretation_by_entry")

CMMR_START_ENTRY = 329

def main():
    # Create output folder
    if not os.path.exists(ENTRY_DIR):
        os.makedirs(ENTRY_DIR)
        print(f"Folder created: {ENTRY_DIR}")

    if not os.path.exists(DB_PATH):
        print(f"Database file not found, please check the path: {DB_PATH}")
        return

    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb.active

    copied = 0
    skipped = 0

    # Iterate through Excel content (starting from the second row)
    for r in range(2, ws.max_row + 1):
        entry_val = ws.cell(r, 1).value
        orig_path = ws.cell(r, 2).value

        # Check entry format
        if not isinstance(entry_val, (int, float)):
            continue
        
        entry_num = int(entry_val)
        
        # Filter entries
        if entry_num < CMMR_START_ENTRY:
            continue
            
        if not orig_path:
            skipped += 1
            continue

        # Combine paths (automatically handle separators)
        src = os.path.join(script_dir, str(orig_path).lstrip('\\/'))
        dst = os.path.join(ENTRY_DIR, f"{entry_num}.txt")

        if os.path.exists(src):
            shutil.copy2(src, dst)
            print(f"Success: Entry {entry_num} -> {os.path.basename(dst)}")
            copied += 1
        else:
            print(f"Warning: Source file not found {src}")
            skipped += 1

    print("-" * 30)
    print(f"Done! {copied} files copied, {skipped} skipped/failed.")

if __name__ == "__main__":
    main()