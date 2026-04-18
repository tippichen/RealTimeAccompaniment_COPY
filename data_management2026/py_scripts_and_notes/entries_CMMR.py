import sys
try:
    import openpyxl
except ImportError:
    import subprocess
    py39 = r"C:\Users\judy\AppData\Local\Programs\Python\Python39\python.exe"
    sys.exit(subprocess.run([py39] + sys.argv).returncode)

import os
import io
import shutil

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

script_dir = os.path.dirname(os.path.abspath(__file__))
WORKSPACE = script_dir
DB_PATH = os.path.join(WORKSPACE, "data_management2026", "database.xlsx")
ENTRY_DIR = os.path.join(WORKSPACE, "data_management2026", "entry")

CMMR_START_ENTRY = 247


def main():
    os.makedirs(ENTRY_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb.active

    copied = 0
    skipped = 0

    for r in range(2, ws.max_row + 1):
        entry_val = ws.cell(r, 1).value
        orig_path = ws.cell(r, 2).value

        if not isinstance(entry_val, (int, float)):
            continue
        entry_num = int(entry_val)
        if entry_num < CMMR_START_ENTRY:
            continue
        if not orig_path:
            print(f"  Entry {entry_num}: no original path, skipping.")
            skipped += 1
            continue

        src = os.path.join(WORKSPACE, str(orig_path))
        dst = os.path.join(ENTRY_DIR, f"{entry_num}.txt")

        if not os.path.exists(src):
            print(f"  Entry {entry_num}: source not found — {orig_path}")
            skipped += 1
            continue

        shutil.copy2(src, dst)
        print(f"  Entry {entry_num}: {orig_path} -> entry\\{entry_num}.txt")
        copied += 1

    print(f"\nDone. Copied {copied} files, skipped {skipped}.")
    print(f"Output directory: {ENTRY_DIR}")


if __name__ == "__main__":
    main()
