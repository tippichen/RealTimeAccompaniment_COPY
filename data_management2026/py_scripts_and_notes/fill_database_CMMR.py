import sys
try:
    import openpyxl
except ImportError:
    import subprocess
    py39 = r"C:\Users\judy\AppData\Local\Programs\Python\Python39\python.exe"
    sys.exit(subprocess.run([py39] + sys.argv).returncode)

import os
import io
import re
import difflib

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

script_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(script_dir)
WORKSPACE = script_dir
RTC = os.path.join(parent_dir, "RealTimeAccompaniment_COPY")
DB_PATH = os.path.join(WORKSPACE, "data_management2026", "database.xlsx")
DATA_CMMR = os.path.join(WORKSPACE, "data_management2026", "CMMR2023")
RTC_LOGS = os.path.join(RTC, "logs")
RTC_CMMR_LOGS = os.path.join(RTC, "CMMR2023", "logs")


def read_first_value(path, prefix):
    try:
        with open(path, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith(prefix):
                    return line[len(prefix):].strip()
    except Exception:
        pass
    return None


def read_path_line(path):
    p = read_first_value(path, "path:")
    if p:
        return p.replace('/', '\\')
    return None


def find_score_path(song_name, fallback=None):
    if song_name and song_name != 'unknown':
        candidate = os.path.join(RTC_LOGS, song_name, "outputscore.txt")
        if os.path.exists(candidate):
            return f"logs\{song_name}\outputscore.txt"
        folders = [f for f in os.listdir(RTC_LOGS)
                   if os.path.isdir(os.path.join(RTC_LOGS, f)) and not f.startswith('.')]
        matches = difflib.get_close_matches(song_name, folders, n=1, cutoff=0.6)
        if matches:
            return f"logs\{matches[0]}\outputscore.txt"
    if fallback:
        return fallback
    return "unknown"


def collect_records():
    records = []
    root_files = sorted(
        f for f in os.listdir(DATA_CMMR)
        if os.path.isfile(os.path.join(DATA_CMMR, f))
        and ('inputinterpretation' in f or 'outputinterpretation' in f)
    )
    for fname in root_files:
        records.append(('root', fname, None))
    logs_dir = os.path.join(DATA_CMMR, "logs")
    for folder in sorted(os.listdir(logs_dir)):
        if folder.lower() == 'old':
            continue
        folder_path = os.path.join(logs_dir, folder)
        if not os.path.isdir(folder_path):
            continue
        for fname in ['inputinterpretation.txt', 'outputinterpretation.txt']:
            if os.path.exists(os.path.join(folder_path, fname)):
                records.append(('logs', fname, folder))
    return records


def make_row(entry_num, file_type, fname, folder):
    if file_type == 'root':
        if '_inputinterpretation' in fname:
            stem = fname.replace('_inputinterpretation.txt', '')
        else:
            stem = fname.replace('_outputinterpretation.txt', '')
        original_path = f"data_management2026\CMMR2023\{fname}"
        raw_txt = os.path.join(RTC, "CMMR2023", f"{stem}.txt")
        raw_src = f"CMMR2023\{stem}.txt" if os.path.exists(raw_txt) else "None"
        gs_path = os.path.join(DATA_CMMR, f"{stem}_guessedscore.txt")
        music_piece = read_first_value(gs_path, "guessed score:") or "unknown"
        fallback = read_path_line(gs_path)
        score_path = find_score_path(music_piece, fallback)
        player = "unknown AI"
        note = f"data_management2026\CMMR2023\{stem}_guessedscore.txt"
    else:
        original_path = f"data_management2026\CMMR2023\logs\{folder}\{fname}"
        if fname == 'inputinterpretation.txt':
            rtc_in = os.path.join(RTC_CMMR_LOGS, folder, "inputmsglog.txt")
            raw_src = (f"CMMR2023\logs\{folder}\inputmsglog.txt"
                       if os.path.exists(rtc_in) else "None")
        else:
            raw_src = "None"
        readme = os.path.join(DATA_CMMR, "logs", folder, "README.md")
        is_sr = folder.startswith('score_recorder')
        if is_sr:
            music_piece = "unknown"
            rtc_out = os.path.join(RTC_CMMR_LOGS, folder, "outputscore.txt")
            score_path = (f"CMMR2023\logs\{folder}\outputscore.txt"
                          if os.path.exists(rtc_out) else "unknown")
            player = "unknown"
            note = (f"data_management2026\CMMR2023\logs\{folder}\README.md"
                    if os.path.exists(readme) else "")
        else:
            song_name = read_first_value(readme, "guessed score:")
            music_piece = song_name or "unknown"
            fallback = read_path_line(readme)
            score_path = find_score_path(song_name, fallback)
            m = re.match(r'^(sec\d+)_', folder)
            player = m.group(1) if m else "unknown"
            note = (f"data_management2026\CMMR2023\logs\{folder}\README.md"
                    if os.path.exists(readme) else "")
    return [
        entry_num,
        original_path,
        raw_src,
        music_piece,
        score_path,
        player,
        "unknown",
        "unknown",
        r"data_management2026\py_scripts_and_notes\align_to_score_CMMR.py",
        None,
        note,
    ]


def main():
    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb.active
    if ws.cell(1, 11).value != 'note':
        ws.cell(1, 11).value = 'note'

    # Remove existing CMMR2023 rows (delete from bottom to top to preserve indices)
    cmmr_rows = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(r, 2).value and 'CMMR2023' in str(ws.cell(r, 2).value)
    ]
    for r in reversed(cmmr_rows):
        ws.delete_rows(r)
    if cmmr_rows:
        print(f"Removed {len(cmmr_rows)} existing CMMR2023 rows.")

    # Find last non-CMMR entry number, then start CMMR from that + 1
    last_entry = 0
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if isinstance(val, (int, float)):
            last_entry = max(last_entry, int(val))
    next_entry = last_entry + 1  # 246 + 1 = 247

    records = collect_records()
    print(f"Found {len(records)} files. Starting from entry {next_entry}.")
    for file_type, fname, folder in records:
        row = make_row(next_entry, file_type, fname, folder)
        ws.append(row)
        print(f"  Entry {next_entry}: {row[1]}")
        next_entry += 1
    try:
        wb.save(DB_PATH)
        print(f"\nDone. Saved to: {DB_PATH}")
        print(f"Added {len(records)} entries.")
    except PermissionError:
        print(f"\n[ERROR] 無法儲存 {DB_PATH}")
        print("請先關閉 Excel 中的 database.xlsx，再重新執行腳本。")


if __name__ == "__main__":
    main()